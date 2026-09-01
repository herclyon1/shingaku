#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""指定校推薦リストから出願候補 README.md と Excel を生成する。

使い方:
    python3 build.py                # 今日を基準日にして再生成
    python3 build.py 2026-10-15     # 基準日を指定

学校から新しいリストが来たら data/ の xlsx を差し替えて再実行するだけ。
本人の条件（出席率・JLPT 点数）は下の ME を書き換える。
"""
import sys, re, json, glob, datetime as dt
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 本人の条件 ───────────────────────────────────────────
ME = {'attendance': 100, 'jlpt': 'N2', 'jlpt_score': 111, 'eju': None}

SRC   = sorted(glob.glob('data/指定校推薦リスト*.xlsx'))[-1]
CAMPUS = {k: v for k, v in json.load(open('data/campus.json', encoding='utf-8')).items()
          if not k.startswith('_')}
FEES   = {k: v for k, v in json.load(open('data/fees.json', encoding='utf-8')).items()
          if not k.startswith('_')}
SHEET = '大学（2026年度）  '
TODAY = dt.date.fromisoformat(sys.argv[1]) if len(sys.argv) > 1 else dt.date.today()

# ── 判定ルール（本人の条件に対する不可・要確認）───────────────
NG = {
 '中日本自動車短期大学（岐阜）': '短期大学（你指定排除）',
 '大阪キリスト教短期大学':      '短期大学（你指定排除）',
 '駒沢女子大学・短期大学':      '女子大学（你指定排除）',
 '大阪女学院大学':             '女子大学（你指定排除）',
 '平安女学院大学':             '女子大学（你指定排除）',
 '近畿大学':      '必须有 EJU（日语300分＋综合科目100分＋总分400分）。不认 JLPT',
 '日本工業大学':   '必须有 EJU（日语230分，还要考数学课程2）。不认 JLPT',
 '神戸学院大学':   '必须有 EJU（日语240分）。不认 JLPT',
 '大阪国際大学':   '必须有 EJU（日语200分）。不认 JLPT',
 '上海大学':      '要求中国籍以外，你不符合',
 '太成学院大学':   f'要求 N2 112 分以上，你 {ME["jlpt_score"]} 分差 1 分（EJU220／JPT525／J.TEST C级600 也可以替代）',
}
WARN = {
 '追手門学院大学（パートナー校制度）':
    '募集要項原文は日语组と英语组それぞれに「いずれかを有する者」＝AND。日语 N3 你够，但**必须另有英语成绩**'
    '（TOEIC800／IELTS5.5／TOEFL iBT74／Duolingo100 之一）。反面：英語能力に応じて最大100%の授業料免除あり',
 '城西国際大学':   '指定校名单上条件栏是空的、日程也未记载（名单标「編集中」）。学费は官网で確認済（138.7万）だが、'
    '出願条件そのものが不明なので判定できない ← 学校の進路担当に確認が必要',
}

S = lambda x: '' if x is None else str(x).strip()

def parse():
    ws = openpyxl.load_workbook(SRC, data_only=True)[SHEET]
    R = [[c.value for c in r] for r in ws.iter_rows()]
    out, i = [], 0
    while i < len(R):
        if S(R[i][5]) == '人数' and S(R[i][6]).startswith('推薦条件'):
            d = R[i + 1]
            rec = {'status': S(d[0]), 'name': S(d[1]), 'quota': S(d[5]), 'cond': S(d[6])}
            j = i + 2
            while j < len(R) and not (S(R[j][5]) == '人数' and S(R[j][6]).startswith('推薦条件')):
                if S(R[j][1]) == '出願期間':
                    n = R[j + 1]
                    for c, k in ((1, '出願期間'), (2, '試験日'), (3, '発表'), (4, '手続締切'), (5, '試験科目')):
                        rec[k] = S(n[c])
                    break
                j += 1
            if rec['name']:
                out.append(rec)
            i = j
            continue
        i += 1
    seen, U = set(), []
    for x in out:
        k = x['name'].split('\n')[0]
        if k in seen:
            continue
        seen.add(k)
        x['short'] = k
        x['gakubu'] = '／'.join(x['name'].split('\n')[1:])
        U.append(x)
    return U

def windows(t):
    """出願期間セルから (開始, 締切) を全部拾い、締切が未来のものだけ返す。"""
    out = []
    for m in re.finditer(
            r'(\d{1,2})[/月](\d{1,2})日?'          # 開始 M月D日
            r'\s*[（(]?[月火水木金土日]?[）)]?'        # 曜日（任意）
            r'(?:\s*\d{1,2}[:：]\d{2})?'           # 時刻（任意）← 名古屋経済のような表記に対応
            r'\s*[～~\-–]\s*'
            r'(?:\d{4}年)?(\d{1,2})[/月](\d{1,2})',  # 終了 M月D日
            t.replace('　', ' ')):
        a, b, c, d = (int(x) for x in m.groups())
        y1 = 2026 if a >= 8 else 2027
        y2 = 2026 if c >= 8 else 2027
        if c < a:
            y2 = y1 + 1
        try:
            o, cl = dt.date(y1, a, b), dt.date(y2, c, d)
            if cl >= TODAY:
                out.append((o, cl))
        except ValueError:
            pass
    return sorted(out)

def quota(q):
    """Excel が数値化した枠（'1.0'）を '1名' に戻す。"""
    q = re.sub(r'^(\d+)\.0$', r'\1名', q.strip())
    return q.replace('\n', '／')

def jcond(c):
    m = re.search(r'(N\s*[12１２][^\n]{0,80})', c)
    return m.group(1)[:80] if m else ('日本語要件の記載なし' if c else '')

def attreq(c):
    m = re.findall(r'出席[率状況][^\n]{0,12}?(\d{2})\s*[％%]', c)
    return int(max(m)) if m else None

# 都道府県 → 地方区分
CHIHO = {
 '大阪府': '近畿', '兵庫県': '近畿', '京都府': '近畿', '奈良県': '近畿',
 '滋賀県': '近畿', '和歌山県': '近畿', '三重県': '近畿',
 '埼玉県': '関東', '千葉県': '関東', '東京都': '関東', '神奈川県': '関東',
 '茨城県': '関東', '栃木県': '関東', '群馬県': '関東',
 '愛知県': '中部', '岐阜県': '中部', '山梨県': '中部', '静岡県': '中部',
 '長野県': '中部', '新潟県': '中部', '富山県': '中部', '石川県': '中部', '福井県': '中部',
 '北海道': '北海道',
}

def regions(loc):
    """campus.json の pref（'京都府 / 北海道' のような複数表記も可）を
    (都道府県リスト, 地方リスト) に正規化する。"""
    raw = loc.get('pref', '')
    prefs = [p.strip() for p in raw.split('/') if p.strip()]
    chiho = []
    for p in prefs:
        c = CHIHO.get(p, '海外' if '中国' in p or '上海' in p else 'その他')
        if c not in chiho:
            chiho.append(c)
    return prefs, chiho


def language_req(x, verdict):
    """日本語要件を実際のハードルで分類する。表示・絞り込みの二次カテゴリ。"""
    n, c = x['short'], x['cond']
    if n in ('近畿大学', '日本工業大学', '神戸学院大学', '大阪国際大学'):
        return 'eju', '只认 EJU'
    if n == '追手門学院大学（パートナー校制度）':
        return 'eng', 'N3 可・但另需英语成绩'
    if n == '京都外国語大学':
        return 'split', '按学科：日本語学科要 N1／グローバル系 N2 可'
    if n == '太成学院大学':
        return 'n2hi', 'N2 需 112 分以上'
    if n == '城西国際大学':
        return 'unknown', '条件未记载'
    if n == '園田学園大学':
        return 'n2', 'N2 可（无 N2 才需 B2 证明）'
    if re.search(r'N\s*3|N３', c):
        return 'n3', 'N3 即可'
    if re.search(r'参照枠|CEFR', c) and not re.search(r'N\s*2|N２', c):
        return 'cefr', 'CEFR 判定'
    if re.search(r'N\s*2|N２|JLPT2', c):
        return 'n2', 'N2 可'
    return 'unknown', '条件未记载'


def exam_kind(sub):
    """試験科目を『その場で日本語を書かされるか』で分類する。
    口頭試問・読み上げは口述なので oral 側。『筆記は実施しない/免除』の明示が最優先。"""
    t = (sub or '').replace('\n', ' ')
    if not t.strip():
        return 'unknown', '未记载'
    # 明示的に筆記なしと書いてある場合を最優先で拾う
    if re.search(r'(筆記試験|日本語\s*[（(]?作文|日本語総合試験)[^。]{0,12}(実施しない|免除)', t):
        return 'oral', '明确不考笔试'
    if re.search(r'小論文|課題論文|筆記試験|独自の日本語能力試験|日本語能力試験を実施', t):
        return 'written', '有笔试／小论文'
    return 'oral', '只有书类＋面试'


def waiver_kind(fee):
    if not fee:
        return 'none', '未查到减免信息'
    if fee.get('hit') is True:
        return 'hit', '减免适用'
    if fee.get('hit') is False:
        return 'no', '第一年无减免'
    return 'maybe', '减免额未定'


def waiver_types(fee):
    """减免の中身を種類別に分ける（入学金/学费打折/出勤率/成绩）。"""
    t = (fee or {}).get('reduction') or ''
    out = []
    if re.search(r'入学金', t):
        out.append('w_adm')
    if re.search(r'学费减|减学费|减 ?\d+%|学费的|学费按|另一套价格', t):
        out.append('w_tui')
    if re.search(r'出勤率', t):
        out.append('w_att')
    if re.search(r'成绩|GPA|学分|优秀', t):
        out.append('w_gpa')
    if re.search(r'收入', t):
        out.append('w_inc')
    return out


# 通学圏（関西内は実際に通えるかで刻む。関西外は引っ越し前提）
ZONE = [
 ('大阪市内', ['大阪市']),
 ('大阪府下', ['堺市', '東大阪', '大東', '八尾', '和泉', '熊取', '茨木', '守口', '枚方', '高槻']),
 ('神戸・阪神', ['神戸市', '芦屋', '尼崎', '西宮', '伊丹', '三木']),
 ('播磨', ['姫路']),
 ('京都', ['京都市', '京都']),
 ('奈良', ['天理', '奈良']),
]

def zone_of(loc):
    prefs = loc.get('pref', '')
    cities = ' '.join(b for _, b in loc.get('campus', []))
    if '中国' in prefs or '上海' in prefs:
        return 'overseas', '海外'
    for key, pats in ZONE:
        if any(p in cities for p in pats):
            return key, key
    if any(p in prefs for p in ('大阪府', '兵庫県', '京都府', '奈良県')):
        return 'kansai_other', '関西・その他'
    return 'far', '要搬家（关西外）'


def quota_band(q):
    z = str.maketrans('０１２３４５６７８９', '0123456789')
    ns = [int(v) for v in re.findall(r'(\d+)\s*名', (q or '').translate(z))]
    if ns:
        t = sum(ns)
        return ('q1', '1 名') if t <= 1 else ('q3', '2〜3 名') if t <= 3 else ('q5', '4 名以上')
    if '若干' in (q or ''):
        return 'qs', '若干名'
    return 'qx', '未注明'


def att_band(a):
    if a is None:
        return 'ax', '未记载'
    return f'a{a}', f'{a}% 以上'


# 指定校リストの誤記を、01_資料索引『大学（2027年度）』の記載で訂正する
DATE_FIX = {
 '名古屋経済大学':
   '⚠ 指定校名单の出願期間欄は「2025年」と誤記。01_資料索引『大学（2027年度）』では '
   '出願開始 2026-09-30／出願終了 2027-02-26 と記載されており、年度は 2026-2027 が正しい。'
   'ページ上の残り日数はこの訂正後の日付で計算している。',
}


# N1 を取ると条件・金額が改善する学校（学費調査と学校資料から判明した分）
N1_GAIN = {
 '関西国際大学':  'N1 取得で給付額に +100,000 円',
 '大阪観光大学':  'N1 なら学費が N2 の 80万 → 70万（年 10万円差）',
 '育英館大学':    'N1 合格で減免が 25% → 50%（年 20万 → 40万）',
 '太成学院大学':  'N1 があれば N2 112 分の不足を回避でき、出願可能になる',
 '京都外国語大学': 'N1 110 分以上なら日本語学科も出願可能になる',
}


def classify(x):
    n = x['short']
    if n in NG:
        return '✕', NG[n]
    if n in WARN:
        return '⚠', WARN[n]
    a = attreq(x['cond'])
    if a is not None and ME['attendance'] < a:
        return '✕', f'要求出勤率 {a}% 以上（你 {ME["attendance"]}%）'
    return '◎', ''

def build():
    U = parse()
    rows = []
    for x in U:
        v, why = classify(x)
        ws_ = windows(x.get('出願期間', ''))
        rows.append({
            'v': v, 'why': why, 'n': x['short'], 'g': x['gakubu'],
            'q': quota(x['quota']),
            'sen': any(k in x['cond'] for k in ('専願', '単願', '必ず入学', '入学意思')),
            'j': jcond(x['cond']), 'a': attreq(x['cond']),
            'sub': x.get('試験科目', ''), 'p': x.get('出願期間', ''),
            't': x.get('試験日', ''), 'r': x.get('発表', ''), 'h': x.get('手続締切', ''),
            'w': ws_, 'cond': x['cond'],
            'loc': CAMPUS.get(x['short'], {}),
            'fee': FEES.get(x['short'], {}),
        })
        rows[-1]['prefs'], rows[-1]['chiho'] = regions(rows[-1]['loc'])
        rows[-1]['req'], rows[-1]['req_label'] = language_req(x, v)
        rows[-1]['exam'], rows[-1]['exam_label'] = exam_kind(x.get('試験科目', ''))
        rows[-1]['wv'], rows[-1]['wv_label'] = waiver_kind(rows[-1]['fee'])
        rows[-1]['zone'], rows[-1]['zone_label'] = zone_of(rows[-1]['loc'])
        rows[-1]['qb'], rows[-1]['qb_label'] = quota_band(rows[-1]['q'])
        rows[-1]['ab'], rows[-1]['ab_label'] = att_band(rows[-1]['a'])
        rows[-1]['wt'] = waiver_types(rows[-1]['fee'])
        rows[-1]['n1gain'] = N1_GAIN.get(x['short'], '')
        rows[-1]['datefix'] = DATE_FIX.get(x['short'], '')
        rows[-1]['reqtop'] = ('eju' if rows[-1]['req'] == 'eju' else
                              'eng' if rows[-1]['req'] == 'eng' else
                              'unknown' if rows[-1]['req'] == 'unknown' else 'jlpt')
        f = rows[-1]['fee']
        amt = f.get('net') or None
        rows[-1]['band'] = ('a' if amt and amt < 1000000 else
                            'b' if amt and amt < 1200000 else
                            'c' if amt else 'x')
    o = {'◎': 0, '⚠': 1, '✕': 2}
    rows.sort(key=lambda r: (o[r['v']], r['w'][0][1] if r['w'] else dt.date(2099, 1, 1)))
    return rows

# ── Excel ────────────────────────────────────────────────
def excel(rows, path='指定校推薦_出願候補.xlsx'):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = '出願候補'
    H = ['判定', '締切まで', '学校名', '学部・学科', '枠', '専願', '日本語条件', '出席率条件',
         '出願期間', '試験日', '合格発表', '入学手続き締切', '試験科目', '判定理由・注意点']
    W = [6, 9, 22, 34, 16, 7, 34, 11, 30, 26, 22, 26, 26, 46]
    thin = Side(style='thin', color='D0CEC7')
    BD = Border(left=thin, right=thin, top=thin, bottom=thin)
    FILL = {'◎': 'E8F0E6', '⚠': 'FBF2DC', '✕': 'F2E6E3'}
    ws.append(H)
    for j, (h, w) in enumerate(zip(H, W), 1):
        c = ws.cell(1, j)
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2C5070')
        c.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[1].height = 30
    for r in rows:
        left = (r['w'][0][1] - TODAY).days if r['w'] else None
        ws.append([r['v'], f'{left}日' if left is not None else '—', r['n'], r['g'], r['q'],
                   '専願' if r['sen'] else '', r['j'],
                   f"{r['a']}%以上" if r['a'] else '記載なし',
                   r['p'], r['t'], r['r'], r['h'], r['sub'], r['why']])
        i = ws.max_row
        for j in range(1, len(H) + 1):
            c = ws.cell(i, j)
            c.border = BD; c.font = Font(size=10)
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.fill = PatternFill('solid', fgColor=FILL[r['v']])
        ws.cell(i, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(i, 1).font = Font(size=13, bold=True)
        ws.cell(i, 2).alignment = Alignment(horizontal='center', vertical='center')
        if left is not None and left <= 30:
            ws.cell(i, 2).font = Font(size=11, bold=True, color='A63A2B')
        if r['sen']:
            ws.cell(i, 6).font = Font(size=10, bold=True, color='A63A2B')
    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = f'A1:N{ws.max_row}'

    w2 = wb.create_sheet('条件全文')
    w2.append(['判定', '学校名', '枠', '推薦条件（原文）'])
    for j, w in enumerate([6, 22, 18, 110], 1):
        c = w2.cell(1, j)
        c.font = Font(bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2C5070')
        c.alignment = Alignment(horizontal='center', vertical='center')
        w2.column_dimensions[get_column_letter(j)].width = w
    for r in rows:
        w2.append([r['v'], r['n'], r['q'], r['cond']])
        i = w2.max_row
        for j in range(1, 5):
            c = w2.cell(i, j)
            c.border = BD; c.font = Font(size=9)
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.fill = PatternFill('solid', fgColor=FILL[r['v']])
    w2.freeze_panes = 'B2'
    wb.save(path)
    return path

# ── README ───────────────────────────────────────────────
def md(rows):
    C = Counter(r['v'] for r in rows)
    ok = [r for r in rows if r['v'] == '◎']
    wn = [r for r in rows if r['v'] == '⚠']
    ng = [r for r in rows if r['v'] == '✕']
    esc = lambda s: (s or '').replace('|', '／').replace('\n', ' ')

    L = []
    L.append('# 指定校推薦 出願候補（2027年4月入学）\n')
    L.append(f'**更新 {TODAY}** ｜ 出席率 {ME["attendance"]}% ｜ JLPT {ME["jlpt"]} {ME["jlpt_score"]}点\n')
    L.append('## 校数\n')
    L.append('| 区分 | 校数 |')
    L.append('|---|---:|')
    L.append(f'| 指定校リスト掲載（大学・2026年度） | **{len(rows)}** |')
    L.append(f'| ◎ この条件で出願可 | **{C["◎"]}** |')
    L.append(f'| ⚠ 追加確認が必要 | {C["⚠"]} |')
    L.append(f'| ✕ 出願不可 | {C["✕"]} |')
    L.append(f'| うち専願（合格＝入学義務） | {sum(1 for r in ok if r["sen"])} / {C["◎"]} |')
    L.append('\n> 別途、専門学校の指定校が 32 校ある（このリストは大学のみ）。\n')

    nxt = [r for r in ok if r['w']]
    if nxt:
        n0 = min(nxt, key=lambda r: r['w'][0][1])
        d0 = (n0['w'][0][1] - TODAY).days
        L.append(f'## 次の締切\n\n**{n0["n"]} — {n0["w"][0][1]:%-m/%-d}（あと {d0} 日）**\n')

    L.append('## ◎ 出願可\n')
    L.append('| 締切 | 残 | 学校 | 枠 | 専願 |')
    L.append('|---|---:|---|---|:-:|')
    for r in ok:
        if r['w']:
            o, c = r['w'][0]
            left = (c - TODAY).days
            dl = f'{c:%-m/%-d}'
            if left <= 14:
                dl = f'**{dl}** 🔴'
            elif left <= 30:
                dl = f'**{dl}** 🟡'
            leftn = str(left)
        else:
            dl, leftn = '—', '—'
        L.append(f'| {dl} | {leftn} | {esc(r["n"])} | {esc(r["q"])} | {"専" if r["sen"] else ""} |')

    L.append('\n## ⚠ 要確認\n')
    for r in wn:
        L.append(f'- **{esc(r["n"])}** — {esc(r["why"])}')
    L.append('\n## ✕ 出願不可\n')
    for r in ng:
        L.append(f'- **{esc(r["n"])}** — {esc(r["why"])}')

    L.append('\n## 各校の詳細\n')
    for r in ok + wn:
        L.append(f'<details><summary><b>{r["v"]} {esc(r["n"])}</b>'
                 + (f' — {esc(r["q"])}' if r['q'] else '')
                 + (' ｜<b>専願</b>' if r['sen'] else '') + '</summary>\n')
        L.append(f'**学部・学科**：{esc(r["g"]) or "—"}  ')
        L.append(f'**日本語条件**：{esc(r["j"]) or "—"}  ')
        L.append(f'**出席率**：{str(r["a"]) + "% 以上" if r["a"] else "記載なし"}  ')
        if r['why']:
            L.append(f'**注意**：{esc(r["why"])}  ')
        L.append('')
        for k, lab in (('p', '出願期間'), ('t', '試験日'), ('r', '発表'),
                       ('h', '入学手続き締切'), ('sub', '試験科目')):
            if r[k]:
                L.append(f'**{lab}**')
                L.append('```')
                L.append(r[k])
                L.append('```')
        L.append('\n</details>\n')

    L.append('---\n')
    L.append('### 前提\n')
    L.append('- 入学は全て **2027年4月**。表中の 2026年10〜12月 は出願・試験日であって入学月ではない。')
    L.append(f'- 出席率 {ME["attendance"]}% のため、最も厳しい 95% 要件も含め全校クリア。この項目は候補を絞る条件にならない。')
    L.append(f'- N2 {ME["jlpt_score"]}点。数値指定があるのは 太成学院 112点（不足）と 京都外国語 100点（クリア）の 2 校のみ。')
    L.append('- **専願＝合格したら入学義務**。併願できないため、複数校への同時出願は原則不可。')
    L.append('- 日付・条件は元シートのセルから機械抽出したもの。**出願前に必ず募集要項の原本で照合すること。**')
    L.append(f'\n出典：`{SRC}`「{SHEET.strip()}」\n')
    L.append('### 更新のしかた\n')
    L.append('学校から新しいリストが来たら `data/` の xlsx を差し替えて:\n')
    L.append('```bash\npython3 build.py && git commit -am "update" && git push\n```')
    return '\n'.join(L) + '\n'

if __name__ == '__main__':
    import os
    import html_gen
    rows = build()
    open('README.md', 'w', encoding='utf-8').write(md(rows))
    os.makedirs('docs', exist_ok=True)
    open('docs/index.html', 'w', encoding='utf-8').write(
        html_gen.render(rows, TODAY, ME, SRC, SHEET))
    p = excel(rows)
    C = Counter(r['v'] for r in rows)
    print(f'基準日 {TODAY} ｜ 全{len(rows)}校  ◎{C["◎"]}  ⚠{C["⚠"]}  ✕{C["✕"]}')
    print(f'  → docs/index.html, README.md, {p}')
